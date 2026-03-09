import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_FOLDER = "output"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def export_excel(results: dict, filename="scan_result", confidence_data=None):
    """
    Export results to Excel with enhanced formatting
    """
    if not results:
        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Results"
        ws['A1'] = "No data to export"
        ws['A1'].font = Font(bold=True)
        filepath = os.path.join(OUTPUT_FOLDER, f"{filename}.xlsx")
        wb.save(filepath)
        return filepath

    # Find grid dimensions
    max_row = 0
    max_col = 0
    
    for cell in results.keys():
        if cell and len(cell) >= 2:
            col = ord(cell[0].upper()) - 64
            try:
                row = int(cell[1:])
                max_col = max(max_col, col)
                max_row = max(max_row, row)
            except:
                continue
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid Scan Results"
    
    # Title with timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    title = ws.cell(row=1, column=1, value=f"Grid Scan Results - {filename} ({timestamp})")
    title.font = Font(bold=True, size=14, color="4B0082")
    title.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col+2)
    
    # Headers
    header_font = Font(bold=True, color="4B0082")
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    ws.cell(row=2, column=1, value="Row/Col").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col+1, value=chr(64 + col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add confidence header if confidence data exists
    if confidence_data:
        conf_cell = ws.cell(row=2, column=max_col+2, value="Confidence")
        conf_cell.font = header_font
        conf_cell.fill = header_fill
        conf_cell.alignment = Alignment(horizontal='center')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Data rows
    for row in range(1, max_row + 1):
        # Row number
        row_cell = ws.cell(row=row+2, column=1, value=row)
        row_cell.font = Font(bold=True)
        row_cell.alignment = Alignment(horizontal='center')
        row_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        row_cell.border = thin_border
        
        # Data cells
        for col in range(1, max_col + 1):
            cell_name = chr(64 + col) + str(row)
            value = results.get(cell_name, "")
            
            cell = ws.cell(row=row+2, column=col+1)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if value:
                try:
                    cell.value = int(value)
                except:
                    try:
                        cell.value = float(value)
                    except:
                        cell.value = value
                
                # Color based on confidence if available
                confidence = 1.0
                if confidence_data and row-1 < len(confidence_data) and col-1 < len(confidence_data[row-1]):
                    confidence = confidence_data[row-1][col-1]
                
                if confidence > 0.8:
                    fill_color = "C6E0B4"  # Light green
                elif confidence > 0.5:
                    fill_color = "FFE699"  # Light yellow
                else:
                    fill_color = "F4B084"  # Light orange
                
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            else:
                cell.value = ""
        
        # Add confidence value if available
        if confidence_data and row-1 < len(confidence_data):
            avg_conf = sum(confidence_data[row-1]) / len(confidence_data[row-1]) if confidence_data[row-1] else 0
            conf_cell = ws.cell(row=row+2, column=max_col+2, value=f"{avg_conf:.1%}")
            conf_cell.border = thin_border
            conf_cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col+1)].width = 15
    
    if confidence_data:
        ws.column_dimensions[get_column_letter(max_col+2)].width = 15
    
    # Create summary sheet
    create_summary_sheet(wb, results, max_row, max_col, confidence_data)
    
    # Save
    filepath = os.path.join(OUTPUT_FOLDER, f"{filename}.xlsx")
    wb.save(filepath)
    return filepath

def create_summary_sheet(wb, results, max_row, max_col, confidence_data):
    """Create summary sheet with statistics"""
    ws_summary = wb.create_sheet("Summary")
    
    # Title
    ws_summary['A1'] = "Scan Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:C1')
    
    # Basic stats
    stats = [
        ("Total Rows:", max_row),
        ("Total Columns:", max_col),
        ("Total Cells:", max_row * max_col),
        ("Filled Cells:", sum(1 for v in results.values() if v)),
        ("Empty Cells:", (max_row * max_col) - sum(1 for v in results.values() if v)),
    ]
    
    for i, (label, value) in enumerate(stats, start=3):
        ws_summary[f'A{i}'] = label
        ws_summary[f'A{i}'].font = Font(bold=True)
        ws_summary[f'B{i}'] = value
    
    # Confidence stats if available
    if confidence_data:
        all_confidences = [c for row in confidence_data for c in row if c > 0]
        if all_confidences:
            ws_summary['A9'] = "Average Confidence:"
            ws_summary['A9'].font = Font(bold=True)
            ws_summary['B9'] = f"{sum(all_confidences) / len(all_confidences):.1%}"
            
            ws_summary['A10'] = "High Confidence (>80%):"
            ws_summary['A10'].font = Font(bold=True)
            ws_summary['B10'] = sum(1 for c in all_confidences if c > 0.8)
            
            ws_summary['A11'] = "Medium Confidence (50-80%):"
            ws_summary['A11'].font = Font(bold=True)
            ws_summary['B11'] = sum(1 for c in all_confidences if 0.5 <= c <= 0.8)
            
            ws_summary['A12'] = "Low Confidence (<50%):"
            ws_summary['A12'].font = Font(bold=True)
            ws_summary['B12'] = sum(1 for c in all_confidences if c < 0.5)
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15