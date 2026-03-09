from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import cv2
import numpy as np
import time
from werkzeug.utils import secure_filename
import logging
from datetime import datetime
import json
import socket

from vision.perspective import correct_perspective
from vision.grid_detector import detect_grid
from vision.cell_extractor import extract_cells
from vision.digit_segmenter import segment_digits
from ocr.ocr_model import recognize_digit
from excel.exporter import export_excel

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Custom JSON encoder for numpy types
class NumpyEncoder(json.JSONEncoder):
    """Special json encoder for numpy types"""
    def default(self, obj):
        if isinstance(obj, (np.int_, np.intc, np.intp, np.int8,
                            np.int16, np.int32, np.int64, np.uint8,
                            np.uint16, np.uint32, np.uint64)):
            return int(obj)
        elif isinstance(obj, (np.float_, np.float16, np.float32, np.float64)):
            return float(obj)
        elif isinstance(obj, (np.ndarray,)):
            return obj.tolist()
        return json.JSONEncoder.default(self, obj)

app = Flask(__name__)
app.json_encoder = NumpyEncoder
CORS(app)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route('/scan', methods=['POST'])
def scan():
    start_time = time.time()
    request_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image uploaded"}), 400
        
        file = request.files['image']
        filename = request.form.get("filename", "").strip()
        if not filename:
            filename = f"scan_result_{int(time.time())}"
        
        # Save image with metadata
        img_path = os.path.join(UPLOAD_FOLDER, f"{request_id}.jpg")
        file.save(img_path)
        
        logger.info(f"\n{'='*60}")
        logger.info(f"📸 Processing: {img_path}")
        
        # Read image with multiple attempts
        image = cv2.imread(img_path)
        if image is None:
            # Try different encoding
            file.seek(0)
            image = cv2.imdecode(np.frombuffer(file.read(), np.uint8), cv2.IMREAD_COLOR)
        
        if image is None:
            return jsonify({"error": "Failed to read image"}), 400
        
        # Preprocess image for better results
        image = preprocess_image(image)
        logger.info(f"📏 Image size: {image.shape}")
        
        # Apply multiple enhancement techniques
        enhanced_images = []
        
        # Try different perspective corrections
        logger.info("🔄 Trying perspective correction method 1...")
        corrected1 = correct_perspective(image)
        if corrected1 is not None:
            logger.info(f"✅ Method 1 successful: {corrected1.shape}")
            enhanced_images.append(corrected1)
        else:
            logger.info("❌ Method 1 failed")
        
        # Try with different parameters
        logger.info("🔄 Trying perspective correction method 2...")
        corrected2 = correct_perspective_advanced(image)
        if corrected2 is not None:
            logger.info(f"✅ Method 2 successful: {corrected2.shape}")
            enhanced_images.append(corrected2)
        else:
            logger.info("❌ Method 2 failed")
        
        if not enhanced_images:
            logger.info("⚠️ No perspective correction worked, using original image")
            enhanced_images = [image]
        
        # Try each enhanced image
        best_result = None
        best_cells = []
        best_score = 0
        
        for idx, enhanced in enumerate(enhanced_images):
            logger.info(f"\n🔄 Trying enhancement method {idx + 1} with image shape {enhanced.shape}")
            
            try:
                # Detect grid
                logger.info(f"  - Detecting grid...")
                grid = detect_grid_enhanced(enhanced)
                if grid is None:
                    logger.info(f"  - Grid detection failed, using enhanced image")
                    grid = enhanced
                else:
                    logger.info(f"  - Grid detected: {grid.shape}")
                
                # Extract cells with multiple methods
                logger.info(f"  - Extracting cells...")
                cells = extract_cells_enhanced(grid)
                
                if cells and len(cells) > 0:
                    rows = len(cells)
                    cols = len(cells[0]) if cells and cells[0] else 0
                    logger.info(f"  - Found {rows} rows, {cols} columns")
                    
                    # Score this result
                    score = rows * cols
                    
                    # Check if cells have content
                    content_score = 0
                    for r in range(min(3, rows)):
                        for c in range(min(3, cols)):
                            if cells[r][c] is not None and cells[r][c].size > 0:
                                if len(cells[r][c].shape) == 3:
                                    gray = cv2.cvtColor(cells[r][c], cv2.COLOR_BGR2GRAY)
                                else:
                                    gray = cells[r][c].copy()
                                if np.mean(gray) > 10:  # Not empty
                                    content_score += 1
                    
                    # Adjust score based on content
                    if content_score > 0:
                        score = score * (1 + content_score / 9)
                    
                    logger.info(f"  - Score: {score:.2f} (content: {content_score}/9)")
                    
                    if score > best_score:
                        best_cells = cells
                        best_result = enhanced
                        best_score = score
                        logger.info(f"  ✅ New best result!")
                else:
                    logger.info(f"  - No cells detected")
                    
            except Exception as e:
                logger.error(f"  ❌ Error in enhancement method {idx + 1}: {str(e)}")
                continue
        
        # If no cells found with enhanced methods, try with original image
        if not best_cells or len(best_cells) == 0:
            logger.info("\n🔄 No cells found with enhanced methods, trying with original image...")
            try:
                cells = extract_cells_enhanced(image)
                if cells and len(cells) > 0:
                    best_cells = cells
                    best_result = image
                    logger.info(f"✅ Found {len(cells)} rows with original image")
            except Exception as e:
                logger.error(f"❌ Error with original image: {str(e)}")
        
        # If still no cells, try smart grid detection
        if not best_cells or len(best_cells) == 0:
            logger.info("\n🔄 Trying smart grid detection...")
            try:
                best_cells = smart_grid_detection(image)
                if best_cells and len(best_cells) > 0:
                    best_result = image
                    logger.info(f"✅ Smart grid detection found {len(best_cells)} rows")
            except Exception as e:
                logger.error(f"❌ Smart grid detection error: {str(e)}")
        
        # Final check
        if not best_cells or len(best_cells) == 0:
            logger.error("❌ No grid cells detected with any method")
            return jsonify({
                "success": False,
                "error": "No grid cells detected. Please ensure the grid is clearly visible.",
                "results": {},
                "grid_data": [],
                "suggestions": [
                    "Ensure good lighting",
                    "Hold camera steady",
                    "Make grid lines visible",
                    "Avoid shadows"
                ]
            })
        
        rows_found = len(best_cells)
        cols_found = len(best_cells[0]) if best_cells and best_cells[0] else 0
        logger.info(f"\n📊 Final result: {rows_found} rows, {cols_found} columns")
        
        results = {}
        grid_data = []
        confidence_data = []
        total_digits = 0
        total_confidence = 0.0
        
        # Process each cell with confidence scoring
        for row_idx, row in enumerate(best_cells):
            row_num = row_idx + 1
            row_data = []
            row_confidence = []
            
            for col_idx, cell in enumerate(row):
                col_num = col_idx + 1
                cell_name = chr(64 + col_num) + str(row_num)
                
                # Enhance cell before processing
                enhanced_cell = enhance_cell_image(cell)
                
                # Get digits with confidence
                digits_info = segment_digits_with_confidence(enhanced_cell)
                
                # Combine digits with confidence
                number = ""
                cell_confidence = 0.0
                digit_count = 0
                
                for digit_img, confidence in digits_info:
                    digit_value, conf = recognize_digit(digit_img, return_confidence=True)
                    if digit_value:
                        number += digit_value
                        cell_confidence += float(conf)  # Convert to Python float
                        digit_count += 1
                        total_digits += 1
                        total_confidence += float(conf)  # Convert to Python float
                
                avg_confidence = cell_confidence / digit_count if digit_count > 0 else 0.0
                
                # Ensure avg_confidence is Python float
                if isinstance(avg_confidence, np.floating):
                    avg_confidence = float(avg_confidence)
                
                results[cell_name] = number
                row_data.append({
                    "cell": cell_name,
                    "value": number,
                    "confidence": round(avg_confidence, 2)
                })
                row_confidence.append(avg_confidence)
                
                if number:
                    logger.debug(f"  Cell {cell_name}: '{number}' (conf: {avg_confidence:.2f})")
            
            grid_data.append(row_data)
            confidence_data.append(row_confidence)
        
        # Export to Excel with formatting
        excel_path = export_excel_enhanced(results, filename, confidence_data)
        
        # Calculate statistics - ensure all values are Python native types
        filled_cells = int(sum(1 for v in results.values() if v))
        processing_time = float(round(time.time() - start_time, 2))
        
        # Convert total_confidence to Python float
        if isinstance(total_confidence, np.floating):
            total_confidence = float(total_confidence)
        
        avg_confidence = float(round(total_confidence / total_digits, 2)) if total_digits > 0 else 0.0
        
        # Convert confidence_data to have Python float values
        converted_confidence_data = []
        for row in confidence_data:
            converted_row = []
            for val in row:
                if isinstance(val, (np.float16, np.float32, np.float64)):
                    converted_row.append(float(val))
                elif isinstance(val, (np.int8, np.int16, np.int32, np.int64)):
                    converted_row.append(int(val))
                else:
                    converted_row.append(val)
            converted_confidence_data.append(converted_row)
        
        # Convert stats to Python native types
        stats = {
            "rows": int(rows_found),
            "columns": int(cols_found),
            "total_cells": int(len(results)),
            "filled_cells": filled_cells,
            "total_digits": int(total_digits),
            "avg_confidence": avg_confidence,
            "processing_time": processing_time
        }
        
        logger.info(f"\n✅ Scan Complete!")
        logger.info(f"📊 Grid: {rows_found} × {cols_found}")
        logger.info(f"📝 Cells: {filled_cells}/{len(results)} filled")
        logger.info(f"🔢 Digits: {total_digits} (avg confidence: {avg_confidence})")
        logger.info(f"⏱️  Time: {processing_time}s")
        logger.info(f"📁 Excel: {excel_path}")
        logger.info(f"{'='*60}\n")
        
        return jsonify({
            "success": True,
            "results": results,
            "grid_data": grid_data,
            "excel_file": f"/download/{os.path.basename(excel_path)}",
            "stats": stats,
            "confidence_data": converted_confidence_data,
            "suggestions": generate_suggestions(filled_cells, len(results), avg_confidence)
        })
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": str(e),
            "success": False,
            "suggestions": ["An error occurred. Please try again."]
        }), 500

def preprocess_image(image):
    """Enhance image for better detection"""
    try:
        # Resize if too large
        h, w = image.shape[:2]
        if max(h, w) > 1200:
            scale = 1200 / max(h, w)
            new_w = int(w * scale)
            new_h = int(h * scale)
            image = cv2.resize(image, (new_w, new_h))
            logger.info(f"  Resized to {new_h}x{new_w}")
        
        # Convert to grayscale if needed
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        # Apply CLAHE for better contrast
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)
        
        # Denoise
        denoised = cv2.fastNlMeansDenoising(enhanced, None, 10, 7, 21)
        
        # Convert back to color if needed
        if len(image.shape) == 3:
            return cv2.cvtColor(denoised, cv2.COLOR_GRAY2BGR)
        return denoised
        
    except Exception as e:
        logger.error(f"Preprocessing error: {e}")
        return image

def correct_perspective_advanced(image):
    """Advanced perspective correction with multiple attempts"""
    try:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5,5), 0)
        
        # Try different threshold methods
        methods = [
            cv2.THRESH_BINARY + cv2.THRESH_OTSU,
            cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU,
            cv2.THRESH_BINARY
        ]
        
        for method in methods:
            try:
                _, thresh = cv2.threshold(blur, 0, 255, method)
                
                # Find contours
                contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                
                if contours:
                    # Find largest contour
                    largest = max(contours, key=cv2.contourArea)
                    
                    # Approximate polygon
                    peri = cv2.arcLength(largest, True)
                    approx = cv2.approxPolyDP(largest, 0.02 * peri, True)
                    
                    if len(approx) == 4:
                        # Apply perspective transform
                        pts = approx.reshape(4, 2)
                        rect = order_points(pts)
                        
                        (tl, tr, br, bl) = rect
                        widthA = np.linalg.norm(br - bl)
                        widthB = np.linalg.norm(tr - tl)
                        heightA = np.linalg.norm(tr - br)
                        heightB = np.linalg.norm(tl - bl)
                        
                        maxWidth = max(int(widthA), int(widthB))
                        maxHeight = max(int(heightA), int(heightB))
                        
                        # Ensure minimum size
                        if maxWidth < 100 or maxHeight < 100:
                            continue
                        
                        dst = np.array([
                            [0, 0],
                            [maxWidth - 1, 0],
                            [maxWidth - 1, maxHeight - 1],
                            [0, maxHeight - 1]
                        ], dtype="float32")
                        
                        M = cv2.getPerspectiveTransform(rect, dst)
                        warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))
                        
                        return warped
            except Exception as e:
                logger.debug(f"  Method failed: {e}")
                continue
        
        return None
        
    except Exception as e:
        logger.error(f"Advanced perspective correction error: {e}")
        return None

def detect_grid_enhanced(image):
    """Enhanced grid detection with multiple methods"""
    try:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Method 1: Adaptive threshold
        thresh1 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                        cv2.THRESH_BINARY_INV, 11, 2)
        
        # Method 2: Otsu's threshold
        _, thresh2 = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        
        # Method 3: Canny edges
        edges = cv2.Canny(gray, 50, 150)
        
        # Combine methods
        combined = cv2.bitwise_or(thresh1, thresh2)
        combined = cv2.bitwise_or(combined, edges)
        
        # Dilate to connect lines
        kernel = np.ones((3,3), np.uint8)
        combined = cv2.dilate(combined, kernel, iterations=1)
        
        # Find grid lines
        lines = cv2.HoughLinesP(combined, 1, np.pi/180, 100, 
                                minLineLength=100, maxLineGap=20)
        
        if lines is not None and len(lines) > 5:
            # Find grid boundaries
            h_lines = []
            v_lines = []
            
            for line in lines:
                x1, y1, x2, y2 = line[0]
                angle = abs(np.arctan2(y2 - y1, x2 - x1) * 180 / np.pi)
                length = np.sqrt((x2-x1)**2 + (y2-y1)**2)
                
                if angle < 20 and length > 50:  # Horizontal
                    h_lines.append(y1)
                elif angle > 70 and length > 50:  # Vertical
                    v_lines.append(x1)
            
            if len(h_lines) > 2 and len(v_lines) > 2:
                # Cluster lines
                h_lines = cluster_positions(h_lines, 30)
                v_lines = cluster_positions(v_lines, 30)
                
                if len(h_lines) >= 2 and len(v_lines) >= 2:
                    # Get boundaries
                    top = min(h_lines)
                    bottom = max(h_lines)
                    left = min(v_lines)
                    right = max(v_lines)
                    
                    # Add padding
                    pad = 20
                    top = max(0, top - pad)
                    bottom = min(image.shape[0], bottom + pad)
                    left = max(0, left - pad)
                    right = min(image.shape[1], right + pad)
                    
                    if bottom > top and right > left:
                        return image[top:bottom, left:right]
        
        # Fallback: use largest contour
        contours, _ = cv2.findContours(combined, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if contours:
            largest = max(contours, key=cv2.contourArea)
            x, y, w, h = cv2.boundingRect(largest)
            
            # Add padding
            pad = 20
            x = max(0, x - pad)
            y = max(0, y - pad)
            w = min(image.shape[1] - x, w + 2*pad)
            h = min(image.shape[0] - y, h + 2*pad)
            
            return image[y:y+h, x:x+w]
        
        return image
        
    except Exception as e:
        logger.error(f"Enhanced grid detection error: {e}")
        return image

def extract_cells_enhanced(image):
    """Enhanced cell extraction with multiple methods"""
    try:
        logger.info(f"  extract_cells_enhanced: image shape {image.shape}")
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Try multiple grid sizes
        grid_sizes = [5, 6, 4, 7, 3, 8]  # Reordered for better chances
        best_cells = []
        best_score = 0
        
        for size in grid_sizes:
            logger.info(f"    Trying grid size {size}x{size}...")
            try:
                # Try line detection first
                cells = extract_cells_by_lines(image, size)
                
                if cells and len(cells) > 0:
                    logger.info(f"      Found {len(cells)} rows with line detection")
                    # Score this grid
                    filled_ratio = count_filled_cells(cells)
                    score = float(filled_ratio * size * size)
                    logger.info(f"      Score: {score:.2f} (filled ratio: {filled_ratio:.2f})")
                    
                    if score > best_score:
                        best_cells = cells
                        best_score = score
                        
                        if score > 10:  # Good enough
                            logger.info(f"      ✅ Acceptable result with score {score}")
                            return best_cells
            except Exception as e:
                logger.error(f"      Error with size {size}: {str(e)}")
                continue
        
        if best_cells:
            logger.info(f"  ✅ Best result: {len(best_cells)} rows with score {best_score}")
            return best_cells
        
        # Fallback to equal division
        logger.info("  Falling back to adaptive division...")
        return divide_into_grid_adaptive(image)
        
    except Exception as e:
        logger.error(f"Enhanced cell extraction error: {e}")
        return []

def extract_cells_by_lines(image, expected_size):
    """Extract cells using line detection"""
    try:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150, apertureSize=3)
        
        # Dilate edges
        kernel = np.ones((2,2), np.uint8)
        edges = cv2.dilate(edges, kernel, iterations=1)
        
        lines = cv2.HoughLinesP(edges, 1, np.pi/180, 80, 
                                minLineLength=40, maxLineGap=15)
        
        if lines is None:
            return []
        
        h_positions = []
        v_positions = []
        
        for line in lines:
            x1, y1, x2, y2 = line[0]
            angle = abs(np.arctan2(y2 - y1, x2 - x1) * 180 / np.pi)
            length = np.sqrt((x2-x1)**2 + (y2-y1)**2)
            
            if angle < 25 and length > 30:  # Horizontal
                h_positions.append(y1)
            elif angle > 65 and length > 30:  # Vertical
                v_positions.append(x1)
        
        if len(h_positions) < 2 or len(v_positions) < 2:
            return []
        
        # Cluster positions
        h_lines = cluster_positions(h_positions, 20)
        v_lines = cluster_positions(v_positions, 20)
        
        if len(h_lines) < 2 or len(v_lines) < 2:
            return []
        
        # Sort
        h_lines = sorted(h_lines)
        v_lines = sorted(v_lines)
        
        # Limit to expected size + 1
        if len(h_lines) > expected_size + 1:
            # Take evenly spaced lines
            indices = np.linspace(0, len(h_lines)-1, expected_size+1, dtype=int)
            h_lines = [h_lines[i] for i in indices]
        
        if len(v_lines) > expected_size + 1:
            indices = np.linspace(0, len(v_lines)-1, expected_size+1, dtype=int)
            v_lines = [v_lines[i] for i in indices]
        
        # Extract cells
        cells = []
        for i in range(len(h_lines) - 1):
            row_cells = []
            for j in range(len(v_lines) - 1):
                y1 = max(0, h_lines[i])
                y2 = min(image.shape[0], h_lines[i + 1])
                x1 = max(0, v_lines[j])
                x2 = min(image.shape[1], v_lines[j + 1])
                
                if y2 > y1 and x2 > x1:
                    cell = image[y1:y2, x1:x2]
                    if cell.size > 0:
                        cell = cv2.resize(cell, (100, 100))
                        row_cells.append(cell)
            
            if row_cells:
                cells.append(row_cells)
        
        return cells
        
    except Exception as e:
        logger.error(f"Line extraction error: {e}")
        return []

def cluster_positions(positions, threshold):
    """Cluster nearby positions"""
    if not positions:
        return []
    
    positions = sorted(positions)
    clusters = []
    current_cluster = [positions[0]]
    
    for pos in positions[1:]:
        if pos - current_cluster[-1] <= threshold:
            current_cluster.append(pos)
        else:
            # Add average of cluster
            clusters.append(int(sum(current_cluster) / len(current_cluster)))
            current_cluster = [pos]
    
    if current_cluster:
        clusters.append(int(sum(current_cluster) / len(current_cluster)))
    
    return clusters

def divide_into_grid_adaptive(image):
    """Adaptive grid division based on image content"""
    try:
        h, w = image.shape[:2]
        
        # Analyze image to determine optimal grid size
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        edges = cv2.Canny(gray, 50, 150)
        
        # Count edge pixels to estimate grid density
        edge_density = float(np.sum(edges > 0)) / (h * w)
        
        # Determine grid size based on density
        if edge_density > 0.3:
            grid_size = 8
        elif edge_density > 0.2:
            grid_size = 6
        elif edge_density > 0.1:
            grid_size = 5
        else:
            grid_size = 4
        
        logger.info(f"    Adaptive grid size: {grid_size}x{grid_size} (density: {edge_density:.2f})")
        
        cell_h = h // grid_size
        cell_w = w // grid_size
        
        cells = []
        for i in range(grid_size):
            row_cells = []
            for j in range(grid_size):
                y1 = i * cell_h
                y2 = (i + 1) * cell_h if i < grid_size - 1 else h
                x1 = j * cell_w
                x2 = (j + 1) * cell_w if j < grid_size - 1 else w
                
                cell = image[y1:y2, x1:x2]
                cell = cv2.resize(cell, (100, 100))
                row_cells.append(cell)
            
            cells.append(row_cells)
        
        return cells
        
    except Exception as e:
        logger.error(f"Adaptive division error: {e}")
        return []

def smart_grid_detection(image):
    """Smart grid detection using multiple techniques"""
    try:
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        # Use morphological operations to find grid
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        closed = cv2.morphologyEx(gray, cv2.MORPH_CLOSE, kernel)
        
        # Find horizontal and vertical lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
        
        horizontal_lines = cv2.morphologyEx(closed, cv2.MORPH_OPEN, horizontal_kernel)
        vertical_lines = cv2.morphologyEx(closed, cv2.MORPH_OPEN, vertical_kernel)
        
        # Combine lines
        grid_lines = cv2.bitwise_or(horizontal_lines, vertical_lines)
        
        # Threshold
        _, thresh = cv2.threshold(grid_lines, 50, 255, cv2.THRESH_BINARY)
        
        # Find contours of grid cells
        contours, _ = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        
        cell_rects = []
        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            area = w * h
            if (w > 30 and h > 30 and 
                w < image.shape[1] * 0.8 and 
                h < image.shape[0] * 0.8 and
                area > 900 and area < 50000):
                cell_rects.append((x, y, w, h))
        
        if len(cell_rects) > 4:
            # Organize into grid
            return organize_cells_advanced(image, cell_rects)
        
        return []
        
    except Exception as e:
        logger.error(f"Smart grid detection error: {e}")
        return []

def organize_cells_advanced(image, rects):
    """Advanced cell organization"""
    if len(rects) < 4:
        return []
    
    # Sort by y-coordinate
    rects.sort(key=lambda r: r[1])
    
    # Group into rows
    rows = []
    current_row = [rects[0]]
    current_y = rects[0][1]
    threshold = image.shape[0] * 0.1
    
    for rect in rects[1:]:
        x, y, w, h = rect
        if abs(y - current_y) <= threshold:
            current_row.append(rect)
        else:
            current_row.sort(key=lambda r: r[0])
            rows.append(current_row)
            current_row = [rect]
            current_y = y
    
    if current_row:
        current_row.sort(key=lambda r: r[0])
        rows.append(current_row)
    
    if len(rows) < 2:
        return []
    
    # Normalize grid size
    expected_cols = max(len(row) for row in rows)
    
    # Extract cell images
    cells = []
    for row in rows:
        row_cells = []
        for x, y, w, h in row:
            # Add padding
            pad = 5
            x1 = max(0, x - pad)
            y1 = max(0, y - pad)
            x2 = min(image.shape[1], x + w + pad)
            y2 = min(image.shape[0], y + h + pad)
            
            cell = image[y1:y2, x1:x2]
            cell = cv2.resize(cell, (100, 100))
            row_cells.append(cell)
        
        # Pad row if needed
        while len(row_cells) < expected_cols:
            row_cells.append(np.zeros((100, 100, 3), dtype=np.uint8))
        
        cells.append(row_cells)
    
    return cells

def enhance_cell_image(cell):
    """Enhance cell image for better digit recognition"""
    try:
        if len(cell.shape) == 3:
            gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
        else:
            gray = cell.copy()
        
        # Apply multiple enhancements
        # 1. Contrast enhancement
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)
        
        # 2. Denoise
        denoised = cv2.fastNlMeansDenoising(enhanced, None, 10, 7, 21)
        
        # 3. Sharpen
        kernel = np.array([[-1,-1,-1],
                           [-1, 9,-1],
                           [-1,-1,-1]])
        sharpened = cv2.filter2D(denoised, -1, kernel)
        
        return sharpened
        
    except Exception as e:
        logger.error(f"Cell enhancement error: {e}")
        return cell

def segment_digits_with_confidence(cell):
    """Segment digits and return with confidence scores"""
    try:
        if cell is None or cell.size == 0:
            return []
        
        # Ensure cell is grayscale
        if len(cell.shape) == 3:
            gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
        else:
            gray = cell.copy()
        
        # Try multiple threshold methods
        digits_info = []
        
        # Method 1: Otsu's threshold
        _, thresh1 = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        digits1 = extract_digits_with_confidence(thresh1)
        digits_info.extend(digits1)
        
        # Method 2: Adaptive threshold
        thresh2 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                        cv2.THRESH_BINARY_INV, 11, 2)
        digits2 = extract_digits_with_confidence(thresh2)
        digits_info.extend(digits2)
        
        # Remove duplicates based on position
        unique_digits = []
        seen_positions = set()
        
        for x, digit_img, confidence in digits_info:
            position_key = round(x / 15)  # Group nearby positions
            if position_key not in seen_positions:
                seen_positions.add(position_key)
                # Convert confidence to Python float
                if isinstance(confidence, np.floating):
                    confidence = float(confidence)
                unique_digits.append((digit_img, confidence))
        
        # Sort by confidence (higher confidence first)
        unique_digits.sort(key=lambda d: d[1], reverse=True)
        
        return unique_digits
        
    except Exception as e:
        logger.error(f"Digit segmentation error: {e}")
        return []

def extract_digits_with_confidence(thresh):
    """Extract digits with position and confidence"""
    digits = []
    
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        area = cv2.contourArea(cnt)
        
        # Filter based on size and aspect ratio
        aspect_ratio = h / w if w > 0 else 0
        if (w > 8 and h > 12 and w < 50 and h < 70 and 
            area > 30 and area < 1500 and
            0.4 < aspect_ratio < 2.5):
            
            # Calculate confidence based on contour properties
            rect_area = w * h
            area_ratio = area / rect_area if rect_area > 0 else 0
            
            # Higher confidence for solid, well-formed digits
            confidence = min(area_ratio, 1.0)
            
            # Extract digit
            pad = 2
            x1 = max(0, x - pad)
            y1 = max(0, y - pad)
            x2 = min(thresh.shape[1], x + w + pad)
            y2 = min(thresh.shape[0], y + h + pad)
            
            digit = thresh[y1:y2, x1:x2]
            
            # Clean and normalize
            cleaned = clean_digit_advanced(digit)
            if cleaned is not None:
                digits.append((x, cleaned, confidence))
    
    # Sort by x position
    digits.sort(key=lambda d: d[0])
    
    return digits

def clean_digit_advanced(digit):
    """Advanced digit cleaning"""
    if digit.size == 0:
        return None
    
    # Find largest contour
    contours, _ = cv2.findContours(digit, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if not contours:
        return None
    
    largest = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(largest)
    
    # Extract just the digit
    digit = digit[y:y+h, x:x+w]
    
    # Create square with padding
    size = max(w, h) + 10
    square = np.zeros((size, size), dtype=np.uint8)
    
    # Center the digit
    y_offset = (size - h) // 2
    x_offset = (size - w) // 2
    square[y_offset:y_offset+h, x_offset:x_offset+w] = digit
    
    # Apply morphological operations to clean
    kernel = np.ones((2,2), np.uint8)
    square = cv2.morphologyEx(square, cv2.MORPH_CLOSE, kernel)
    square = cv2.morphologyEx(square, cv2.MORPH_OPEN, kernel)
    
    # Resize to 28x28
    final = cv2.resize(square, (28, 28))
    
    return final

def count_filled_cells(cells):
    """Count cells that likely contain digits"""
    filled = 0
    total = 0
    
    for row in cells:
        for cell in row:
            total += 1
            if cell is not None and cell.size > 0:
                if len(cell.shape) == 3:
                    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
                else:
                    gray = cell.copy()
                
                # Count non-zero pixels
                non_zero = np.count_nonzero(gray > 50)
                if non_zero > 50:  # Threshold for containing content
                    filled += 1
    
    return float(filled) / float(total) if total > 0 else 0.0

def generate_suggestions(filled_cells, total_cells, confidence):
    """Generate helpful suggestions based on results"""
    suggestions = []
    
    if filled_cells == 0:
        suggestions.append("No digits detected. Try:")
        suggestions.append("• Ensure good lighting")
        suggestions.append("• Make digits larger and clearer")
        suggestions.append("• Avoid shadows on the grid")
    elif filled_cells < total_cells * 0.5:
        suggestions.append("Only half the cells detected. Tips:")
        suggestions.append("• Write digits more clearly")
        suggestions.append("• Ensure grid lines are visible")
        suggestions.append("• Hold camera steady")
    elif confidence < 0.7:
        suggestions.append("Low recognition confidence. Suggestions:")
        suggestions.append("• Use darker ink")
        suggestions.append("• Avoid overlapping digits")
        suggestions.append("• Keep digits within cells")
    else:
        suggestions.append("Good scan! For best results:")
        suggestions.append("• Keep consistent lighting")
        suggestions.append("• Use high contrast pens")
    
    return suggestions

def export_excel_enhanced(results, filename, confidence_data=None):
    """Enhanced Excel export with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import os
    
    OUTPUT_FOLDER = "output"
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    if not results:
        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Results"
        ws['A1'] = "No data to export"
        ws['A1'].font = Font(bold=True)
        filepath = os.path.join(OUTPUT_FOLDER, f"{filename}.xlsx")
        wb.save(filepath)
        return filepath
    
    # Find grid dimensions
    max_row = 0
    max_col = 0
    
    for cell in results.keys():
        if cell and len(cell) >= 2:
            col = ord(cell[0].upper()) - 64
            try:
                row = int(cell[1:])
                max_col = max(max_col, col)
                max_row = max(max_row, row)
            except:
                continue
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid Scan Results"
    
    # Title with timestamp
    title = ws.cell(row=1, column=1, value=f"Grid Scan Results - {filename} ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    title.font = Font(bold=True, size=14, color="4B0082")
    title.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col+2)
    
    # Headers with style
    header_font = Font(bold=True, color="4B0082")
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    ws.cell(row=2, column=1, value="Row/Col").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col+1, value=chr(64 + col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add confidence header if available
    if confidence_data:
        conf_cell = ws.cell(row=2, column=max_col+2, value="Confidence")
        conf_cell.font = header_font
        conf_cell.fill = header_fill
        conf_cell.alignment = Alignment(horizontal='center')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Data rows
    for row in range(1, max_row + 1):
        # Row number
        row_cell = ws.cell(row=row+2, column=1, value=row)
        row_cell.font = Font(bold=True)
        row_cell.alignment = Alignment(horizontal='center')
        row_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        row_cell.border = thin_border
        
        # Data cells
        for col in range(1, max_col + 1):
            cell_name = chr(64 + col) + str(row)
            value = results.get(cell_name, "")
            
            cell = ws.cell(row=row+2, column=col+1)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if value:
                try:
                    cell.value = int(value)
                except:
                    try:
                        cell.value = float(value)
                    except:
                        cell.value = value
                
                # Color based on confidence
                confidence = 1.0
                if confidence_data and row-1 < len(confidence_data) and col-1 < len(confidence_data[row-1]):
                    confidence = confidence_data[row-1][col-1]
                
                if confidence > 0.8:
                    fill_color = "C6E0B4"  # Light green
                elif confidence > 0.5:
                    fill_color = "FFE699"  # Light yellow
                else:
                    fill_color = "F4B084"  # Light orange
                
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            else:
                cell.value = ""
        
        # Add confidence value if available
        if confidence_data and row-1 < len(confidence_data):
            avg_conf = sum(confidence_data[row-1]) / len(confidence_data[row-1]) if confidence_data[row-1] else 0
            conf_cell = ws.cell(row=row+2, column=max_col+2, value=f"{avg_conf:.1%}")
            conf_cell.border = thin_border
            conf_cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col+1)].width = 15
    
    if confidence_data:
        ws.column_dimensions[get_column_letter(max_col+2)].width = 15
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary['A1'] = "Scan Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:C1')
    
    ws_summary['A3'] = "Total Rows:"
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['B3'] = max_row
    
    ws_summary['A4'] = "Total Columns:"
    ws_summary['A4'].font = Font(bold=True)
    ws_summary['B4'] = max_col
    
    ws_summary['A5'] = "Total Cells:"
    ws_summary['A5'].font = Font(bold=True)
    ws_summary['B5'] = max_row * max_col
    
    ws_summary['A6'] = "Filled Cells:"
    ws_summary['A6'].font = Font(bold=True)
    ws_summary['B6'] = sum(1 for v in results.values() if v)
    
    ws_summary['A7'] = "Empty Cells:"
    ws_summary['A7'].font = Font(bold=True)
    ws_summary['B7'] = (max_row * max_col) - sum(1 for v in results.values() if v)
    
    if confidence_data:
        all_confidences = [c for row in confidence_data for c in row if c > 0]
        if all_confidences:
            ws_summary['A9'] = "Average Confidence:"
            ws_summary['A9'].font = Font(bold=True)
            ws_summary['B9'] = f"{sum(all_confidences) / len(all_confidences):.1%}"
            
            ws_summary['A10'] = "High Confidence (>80%):"
            ws_summary['A10'].font = Font(bold=True)
            ws_summary['B10'] = sum(1 for c in all_confidences if c > 0.8)
            
            ws_summary['A11'] = "Medium Confidence (50-80%):"
            ws_summary['A11'].font = Font(bold=True)
            ws_summary['B11'] = sum(1 for c in all_confidences if 0.5 <= c <= 0.8)
            
            ws_summary['A12'] = "Low Confidence (<50%):"
            ws_summary['A12'].font = Font(bold=True)
            ws_summary['B12'] = sum(1 for c in all_confidences if c < 0.5)
    
    # Adjust summary column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    # Save
    filepath = os.path.join(OUTPUT_FOLDER, f"{filename}.xlsx")
    wb.save(filepath)
    return filepath

def order_points(pts):
    """Order points for perspective transform"""
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect

def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip

LOCAL_IP = get_local_ip()

@app.route('/download/<filename>')
def download_file(filename):
    try:
        filepath = os.path.join(OUTPUT_FOLDER, filename)
        if not os.path.exists(filepath):
            return jsonify({"error": "File not found"}), 404
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 404

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat()
    })

@app.route('/mobile-info', methods=['GET'])
def mobile_info():
    return jsonify({
        "local_ip": LOCAL_IP,
        "port": 5000,
        "url": f"http://{LOCAL_IP}:5000",
        "status": "online"
    })

if __name__ == "__main__":
    print("\n" + "="*60)
    print(" Advanced Grid Scanner Server Starting...")
    print("="*60)
    print(f" Mobile URL: http://{LOCAL_IP}:5000")
    print(f" Health check: http://{LOCAL_IP}:5000/health")
    print("="*60 + "\n")
    app.run(host="0.0.0.0", port=5000, debug=True, threaded=True)